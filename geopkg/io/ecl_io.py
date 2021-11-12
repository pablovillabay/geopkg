import os
import ntpath
import pandas as pd
import numpy as np
import datetime
from struct import unpack
from numpy import float32, uint32, dtype, float64, array
import openpyxl


def make_smspec_list(path, key):
    
    """

    Iterate over all files in a given path and fetch SMSPEC files which comply with a certain string filter

    Parameters:
        path (string): path to the cases folder
        key (string): filtering string, the file name MUST contain the string in order to be returned in the list

    Returns:
        smspec_list (list:string): list with all SMSPEC file names, complete paths

    """

    smspec_list = []
    for root, dirs, files in os.walk(path):
        for file in files:
            if '.SMSPEC' in file and key in file:  # 'key' is the filtering string to extract relevant cases from path
                smspec_list.append(os.path.join(root, file))
    return smspec_list


def read_summary(filename):
    
    """

    Read binary summary file file and unpack the information

    Parameters:
        filename (string): file name (complete path) of summary Eclipse file

    Returns:
        data (dict): 

    """

    bsdict, data = {b'REAL': (4, 'f', float32), b'INTE': (4, 'I', uint32), b'DOUB': (8, 'd', float64),
                    b'LOGI': (4, 'I', uint32), b'CHAR': (8, 's', dtype('a8'))}, {}

    with open(filename, 'rb+') as f:
        fr = f.read
        while 1:
            head = fr(20)
            if len(head) == 0:
                break
            else:
                _, name, datalen, type = unpack('>L8sL4s', head)
                name = name.strip()
                if datalen == 0:
                    break
                else:
                    buf = unpack('>L', fr(8)[4:])[0]
                    bsize, fmt, npfmt = bsdict[type]
                    tchar = (1 if type != b'CHAR' else 0)
                    if buf == (bsize * datalen):
                        str1 = ('>%d%s' % (datalen, fmt) if tchar else ('>' + '8s' * datalen))
                        rd1 = fr(bsize * datalen)
                        m = unpack(str1, rd1)
                    else:
                        m = []
                        while len(m) < datalen:
                            m += unpack(('>%d%s' % (buf / bsize, fmt) if tchar else ('>' + '8s' * int(buf / bsize))),
                                        fr(buf))
                            buf = (unpack('>L', fr(8)[4:])[0] if len(m) < datalen else buf)
                    fr(4)
                    m = array(m, dtype=npfmt)
                    if name not in data:
                        data[name] = [m]
                    else:
                        data[name] += [m]
    return data 


def get_case_start(smspec):
    
    """

    Get the start date from a simulation case read in its SMSPEC file, return as datetime object

    Parameters:
        filename (string): file name (complete path) of smspec Eclipse file

    Returns:
        start_date (datetime): date of start of simulation

    """
    

    try:
        start_day = smspec[b'STARTDAT'][0][0]
    except:
        print('Error getting case start date')

    start_month = smspec[b'STARTDAT'][0][1]
    start_year = smspec[b'STARTDAT'][0][2]
    # A combination of a date and a time. Attributes: year, month, day, hour, minute, second, microsecond, and tzinfo.
    start_date = datetime.datetime(start_year, start_month, start_day)
    return start_date


def new_wells_info(df):
    
    """

    Check for new wells in the case dataframe and return basic info only from new wells; it returns a dict in 
    which keys are new wells and values are lists of pairs of first non-zero index and functionality to get the 
    first date of production afterwards and keep the functionality

    Parameters:
        df (pd.DataFrame): case dataframe

    Returns:
        new_wells (dict): new wells, index of first production/injection and functionality (producer, injector,...)

    """
    

    new_wells = {}
    wells_dict = {}
    vector_list = ['WWPT', 'WWIT', 'WGPT', 'WGIT', 'WLPT']  # all relevant well prod/inj vectors
    for col in df.columns:
        well = str(col[1])[2:-1]  # clean well name in df multi-labeled column
        if well not in wells_dict.keys() and '+' not in well and str(col[0])[2] == 'W':  # if it is a new well
            wells_dict[well] = 0  # add well to dict and assign zero
        if well in wells_dict.keys() and str(col[0])[2:-1] in vector_list:
            wells_dict[well] += df.loc[0, col]  # accumulate intial values from well vectors for each well in the dict
    for well in wells_dict.keys():
        if wells_dict[well] == 0:
            new_wells[well] = []  # only new wells have 0 in all vectors at the start date of the case, put empty lists
    del wells_dict  # save a bit of memory

    for vector in vector_list:  # loop over vector wells
        for well in new_wells.keys():  # loop over new wells only
            for col in df.columns:  # loop over case df columns
                if well in str(col[1]) and vector in str(col[0]):  # check if column is a well vector for this well
                    if df[col].sum != 0:  # check there is prod/inj data in the column
                        if df[col].ne(0).idxmax() > 0:  # initial date of production cannot be the initial date of case
                            if len(new_wells[well]) == 0:  # first assignment to the specific well
                                new_wells[well].append(df[col].ne(0).idxmax())  # index of first non-zero prod/inj data
                                if 'P' in vector: new_wells[well].append('Producer')  # add well's functionality
                                if 'I' in vector: new_wells[well].append('Injector')
                            else:
                                if new_wells[well][0] < df[col].ne(0).idxmax():
                                    new_wells[well][0] = df[col].ne(0).idxmax()
                                    if 'P' in vector: new_wells[well][1] = 'Producer'
                                    if 'I' in vector: new_wells[well][1] = 'Injector'

    i = 0
    while i < len(new_wells):
        well = list(new_wells.keys())[i]
        if len(new_wells[list(new_wells.keys())[i]]) == 0: # it means it is a new well which never goes active
            new_wells.pop(well, 'not found')
        else:
            i += 1

    return new_wells


def map_full_well_names(dir, case_name):
    
    """

    Define a dict map between short 8char names in Eclipse and full Petrel names as specified in the SCH file, it works 
    by parsing the SCH file line by line, find the begining of the WELSPECS section and read first and last words

    Parameters:
        dir (string): path to case folder
        case_name (string): case name

    Returns:
        well_mapper (dict): mapper between Eclipse namesand full Petrel names as seen in the SCH file

    """

    sch_file = dir + '/' + case_name + '_SCH.INC'  # build the file name to point to the SCH file
    dat_file = dir + '/' + case_name + '.DATA'  # build the file name to point to the DATA file
    no_wells = 0
    with open(dat_file, 'r') as file:
            while True:
                line = file.readline()
                if 'WELLDIMS' in line:  # header of the WELSPECS section
                    line = file.readline()
                    no_wells = int(line.split()[0])
                    break

    well_mapper = {}  # initialize the new dict [old_wellname]: 'new_wellname'
    with open(sch_file, 'r') as file:
        while len(well_mapper) < no_wells:
            line = file.readline()
            if 'WELSPECS' in line:  # header of the WELSPECS section
                line = file.readline()
                if line[:2] == '--':  # all relevant lines start as Eclipse comment
                    well_mapper[line.split()[0][3:-1]] = line.split()[-1][1:-1]
                    line = file.readline()  # it has a final empty comment line, no harm

    return well_mapper


def annualize_profile(df, vectors=['FOPT', 'FLPT', 'FWPT', 'FWIT', 'FGPT', 'FGIT']):
    
    """

    Get an anualized dataframe from relevant field vectors alone; include first and last dates appropiately

    Parameters:
        df (pd.DataFrame): case dataframe
        vectors (list:string): list of interesting vector to be retrieved to the annual profile

    Returns:
        anual_df (pd.DataFrame): annualized profile with relevant vectors only

    """

    anual_dates = []
    for year in range(df.index[0].year, df.index[-1].year + 1): # iterate over possible years of production
        anual_dates.append(pd.Timestamp(str(year)+'-01-01')) # all timesteps but the last must be 01-01-YEAR
    if df.index[-1] > anual_dates[-1]: anual_dates.append(df.index[-1]) # if the case terminates later than 01-Jan, add the precise date

    for vector in vectors:
        if vector not in df.columns:
            vectors.remove(vector) # keep only the interesting vectors that are effectively in the case df
    df = df[vectors].reset_index().drop_duplicates(subset='index', keep='last').set_index('index') # drop duplicate indexes

    anual_df = pd.DataFrame(index=anual_dates, columns=vectors) # initialize empty df with dates as idex and the relevat columns

    # in each row the 01-Jan-YEAR stores the forthcoming production of the year as the incremental FOPT, FGPT,...
    for idx in range(len(anual_df)-1):
        if idx == 0:
            anual_df.loc[anual_df.index[idx], :] = df.loc[anual_df.index[idx+1], :] - df.loc[df.index[idx], :]
        else:
            anual_df.loc[anual_df.index[idx], :] = df.loc[anual_df.index[idx+1], :] - df.loc[anual_df.index[idx], :]
    # if the case terminates later than 01-Jan, add the last year's production
    anual_df.loc[anual_df.index[-1], :] = df.loc[df.index[-1], :] - df.loc[anual_df.index[-2], :]

    return anual_df


def read_case_profiles(path, key, well_info_path=None, vectors=['FOPT', 'FLPT', 'FWPT', 'FWIT', 'FGPT', 'FGIT']):
    
    """

    General function to read all cases in a folder. It returns a dict keyed with case names and valued with a list
    of well info, case df and summary production df

    Parameters:
        path (string): path to cases folder
        key (string): string to filter case names 
        well_info_path (string): path to basic well information file (coordinates, ...)
        vectors (list:string): list of interesting vector to be retrieved to the annual profile

    Returns:
        cases_dict (dict): dict with case names as keys and well info and cases data frames as values

    """

    if well_info_path is not None:
        extra_info = pd.read_csv(well_info_path, sep=',', index_col=0, skipinitialspace=True)
        extra_info.columns = [col[1:] for col in extra_info.columns]
    smspec_list = make_smspec_list(path, key)  # make the cases files list
    cases_dict = {}  # initialize the return dict
    for file_name in smspec_list:
        print(file_name)
        smspec = read_summary(file_name)  # read the smspec
        start_date = get_case_start(smspec)  # get case start date

        def list_strip(lst):
            return list(map(lambda x: x.strip(), smspec[lst][0]))

        try:
            kw, kn, un = list(map(list_strip, (b'KEYWORDS', b'WGNAMES', b'UNITS')))
            cols = [kw, kn, smspec[b'NUMS'][0], un]  # 'NUMS' = CELL | REGION NUMBER
            case_name = ntpath.basename(file_name).split('.')[0]
            dir = os.path.split(file_name)[0]
            try:
                if os.path.isfile(dir + '/' + case_name + '.UNSMRY'):
                    # get case production into a dataframe with SMSPEC column labels
                    unsmry = read_summary(dir + '\\' + case_name + '.UNSMRY')  # Unified summary file
                    df = pd.DataFrame(array(unsmry[b'PARAMS']), columns=cols)
                else:
                    df_list = []
                    for root, dirs, files in os.walk(dir):
                        for file in files:
                            file_split = file.split(sep='.')
                            if file_split[-1][0] == 'S' and file_split[-1][1:].isdigit():
                                summfile = read_summary(root+'\\'+file)
                                df_list.append(pd.DataFrame(np.array(summfile[b'PARAMS']), columns=cols))
                    df = pd.concat(df_list, axis=0, ignore_index=True)
            except:
                continue
                        
            new_wells_df = new_wells_info(df)

            # converting days from start_date to year-month-day format
            # b'TIME' contains time in days from the start_date
            df.index = [start_date + datetime.timedelta(days=int(i[b'TIME'])) for _, i in df.iterrows()]
            df.rename_axis('Date', inplace=True)

            # we want to define simplified column name for out DataFrame using only the Eclipse keywords as headings
            col_list = []
            new_cols = []

            for col in df.columns:
                if b'W' in col[0] in col:  # keep YEAR, TIME and all well vectors
                    col_list.append(col)
                    new_cols.append((str(col[0])[2:][:-1], str(col[1])[2:][:-1]))  # clean mnemonics from column original names

            df = df[col_list]  # filter only the interesting columns
            df.columns = pd.MultiIndex.from_tuples([tup[::-1] for tup in new_cols], names=['Well','Vector'])
            
            df_list = []
            for col in list(df.columns.levels[0]):
                if '+:+' not in col and 'FIELD' not in col:
                    dfwell = df[[col]].copy()
                    dfwell.columns = dfwell.columns.droplevel()
                    dfwell['Wellname'] = col
                    dfwell['Date'] = dfwell.index
                    dfwell.reset_index(inplace=True, drop=True)
                    df_list.append(dfwell)
            df = pd.concat(df_list, axis=0)
            df.set_index(['Wellname', 'Date'], inplace=True)

            well_mapper = map_full_well_names(dir, case_name)  # get a map from 8char to full well names in SCH file

            for well in new_wells_df:
                new_wells_df[well][0] = df.index[new_wells_df[well][0]]  # map wells indexes to dates
            # return the new wells info dict as a clean dataframe
            new_wells_df = pd.DataFrame(list(new_wells_df.values()), index=list(new_wells_df.keys()),
                                        columns=['init_date', 'function'])
            new_wells_df.index.map(mapper=well_mapper)
            if well_info_path is not None: 
                new_wells_df = pd.merge(new_wells_df, extra_info, left_index=True, right_index=True, how='left')
            cases_dict[case_name] = [new_wells_df, df]  # store all case info in the return dict
            del df

        except Exception as e:
            print('Problem when importing case: '+ case_name)
            print(str(e))

    return cases_dict


def export_to_TI(ipt_xls, opt_xls, cases_dict, yrs_idx=10, cncpt_idx=2):

    """

    Map Eclipse case results to TI Excel formats

    Parameters:
        ipt_xls (string): path to empty TI Excel format
        opt_xls (string): path to save created TI Excel sheets
        cases_dict (dict): dict with case names as keys and well info and cases data frames as values
        yrs_idx (int): column index of the first year in th TI format
        cncpt_idx (int): row index for the first concept in the TI format      

    """

    items = {'1A. Sales Crude (Mbbl)': 'FOPT',
             '1B. Sales Condensate (Mbbl)': '',
             '1D. Sales NGL (Mbbl)': '',
             '1D. Sales Gas (Bscf)': 'FGPT',
             '1E. TOTAL SALES (MBOE)': '',
             '1F. Exploration - G&A (M$)': '',
             '1G. Exploration - G&G (includes seismic) (M$)': '',
             '1H. Exploration - Explo Drilling (M$)': '',
             '1I. Exploration - Appraisal Drilling (M$)': '',
             '1J. TOTAL EXPLORATION CAPEX (M$)': '',
             '1K. Development - G&A (M$)': '',
             '1L. Development - G&G (includes seismic) (M$)': '',
             '1M. Development - Drilling Producers (M$)': '',
             '1N. Development - Drilling Injectors/Disposal/Stratigraph. (M$)': '',
             '1O. Development - Recompl. Explo/Appr. wells to producers (M$)': '',
             '1S. Development - Workover and Well Intervention (M$)': '',
             '1P. Development - Production Facilities (M$)': '',
             '1Q. Development - Flowlines and Gathering (M$)': '',
             '1R. Development - Export Pipelines (M$)': '',
             '1T. Development - Others (M$)': '',
             '1U. TOTAL DEVELOPMENT CAPEX (M$)': '',
             '1V. Abandonment - Wells (M$)': '',
             '1W. Abandonment - Facilities and Others (M$)': '',
             '1X. TOTAL ABANDONMENT EXPENDITURE (M$)': '',
             '1Y. Costs- Operating personnel (M$)': '',
             '1Z. Costs - Inspection and maintenance (M$)': '',
             '1ZA. Costs - Well costs (M$)': '',
             '1ZB. Costs - Transport (M$)': '',
             '1ZC. Costs - Other (M$)': '',
             '1ZD. TOTAL OPEX (M$)': '',
             '1ZE. Other Repsol Costs non shareable w partners - net figures (M$)': '',
             '4A. Reserves incorporation to PD Liquids (Mbbl)': '',
             '4B. Reserves incorporation to PD Gas (Bcf)': '',
             '4C. RESERVES INCORPORATION TO PD (Mboe)': '',
             '4E. Reserves incorporation to Total Proved Liquids (Mbbl)': '',
             '4F. Reserves incorporation to Total Proved Gas (Bcf)': '',
             '4G. RESERVES INCORPORATION TO PT (Mboe)': ''}

    try:
        wb = openpyxl.load_workbook(ipt_xls, data_only=True)
        ws = wb.worksheets[0]
        df_test = cases_dict[list(cases_dict.keys())[0]][1].copy()
        df_test.index = df_test.index.year
        year_row = ws[yrs_idx]
        years_list = []
        for cell in year_row: years_list.append(cell.value)
        first_column = years_list.index(df_test.index.min()) + 1
        if 'Crude' not in ws.cell(row=yrs_idx+1, column=cncpt_idx).value:
            print('There was a problem reading the Excel base format, check the years row index and concepts columns index')


    except:
        print('There was a problem reading the Excel base format, check the years row index and concepts columns index')

    for case in cases_dict:

        wb = openpyxl.load_workbook(ipt_xls, data_only=True)
        ws = wb.worksheets[0]
        profile = cases_dict[case][1]
        profile.index = profile.index.year

        for row_idx in range(200):
            concept = ws.cell(row=row_idx+1, column=cncpt_idx).value
            if concept in items.keys() and items[concept] in profile.columns:
                col_counter = 0
                for idx, row in profile.iterrows():
                    ws.cell(row=row_idx+1, column=first_column+col_counter).value = row[items[concept]]
                    col_counter += 1

        wb.save(opt_xls+'\ExportTI_'+case+'.xlsx')
        wb.close()

        book = openpyxl.load_workbook(opt_xls+'\ExportTI_'+case+'.xlsx')
        writer = pd.ExcelWriter(opt_xls+'\ExportTI_'+case+'.xlsx', engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        cases_dict[case][0].to_excel(writer, "WellActivity")
        writer.save()


def import_prod_data(volfile):
    
    """

    Import production data from *.vol formats (OFM or Petrel typical export)

    Parameters:
        path (string): path to cases folder

    Returns:
        cases_dict (dict): dict with case names as keys and well info and cases data frames as values

    """

    print('Reading production data from: ', volfile)

    data = []
    well = ''
    skip = -999

    with open(volfile) as f:
        section = 'info'
        for i, l in enumerate(f):  # read through the lines in the file (i counts and l stores each line)
            if l == '*FIELD\n' and section == 'info': 
                print('Field data')

            elif l == '*DAILY\n' and section == 'info':
                print('Daily production data reported')

            elif 'MISSING_VALUE' in l and section == 'info':
                skip = int(l.split()[-1])
                print('Missing data reported as: ', str(skip))

            elif l.split()[0] == '*DAY' and section == 'info':
                headers = ['Well']
                [headers.append(elem[1:]) for elem in l.split()]
                print('Reported vectors:')
                print(headers)

            elif l.split()[0] == '*NAME':
                section = 'data'
                well = l.split()[-1]
                print('Reading data for well: ', well)

            else:
                if section == 'data':
                    new_line = [well]
                    [new_line.append(elem) for elem in l.split()]
                    data.append(new_line)

    voldata = pd.DataFrame.from_records(data, columns=headers)
    voldata[headers[1:]] = voldata[headers[1:]].apply(pd.to_numeric)
    voldata.replace(to_replace=skip, value=np.nan, inplace=True)  
    voldata['Date'] = pd.to_datetime(voldata[['YEAR', 'MONTH', 'DAY']])
    for col in ['HOUR', 'MINUTE', 'SECOND', 'YEAR', 'MONTH', 'DAY']:
        if col in voldata.columns:
            voldata.drop(col, axis=1, inplace=True)
    voldata.set_index(['Well', 'Date'], inplace=True, drop=True)
    return voldata